import io
import os
import re
import json
import tempfile
import threading
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
from collections import defaultdict, Counter

from PIL import Image, ImageOps
import pytesseract
import torch
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, PlainTextResponse
from transformers import AutoProcessor, LayoutLMv3ForTokenClassification
from openpyxl import Workbook, load_workbook

# --- .env auto-loader (optional) ---
try:
    from dotenv import load_dotenv, find_dotenv  # pip install python-dotenv
    load_dotenv(find_dotenv(), override=False)
except Exception:
    pass

# ===================== Config =====================
MODEL_DIR = os.environ.get("MODEL_DIR", "best_model")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "data/invoices.xlsx")
MAX_LENGTH = int(os.environ.get("MAX_LENGTH", "512"))
ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS", "*")

# Sinks: excel | gsheets | both
WRITE_SINK = os.environ.get("WRITE_SINK", "excel").lower().strip()

# Google Sheets config
GSHEETS_ID = os.environ.get("GSHEETS_ID", "").strip()
GSHEETS_URL = os.environ.get("GSHEETS_URL", "").strip()
if not GSHEETS_ID and GSHEETS_URL:
    m = re.search(r"/d/([a-zA-Z0-9-_]+)/", GSHEETS_URL)
    if m:
        GSHEETS_ID = m.group(1)

GOOGLE_APPLICATION_CREDENTIALS = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
GOOGLE_CRED_JSON = os.environ.get("GOOGLE_CRED_JSON", "").strip()

GSHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

app = FastAPI(title="Invoice Extractor (LayoutLMv3)", version="1.3.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[ALLOWED_ORIGINS] if ALLOWED_ORIGINS != "*" else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logger = logging.getLogger("uvicorn.error")

# Try to import gspread; keep optional
try:
    import gspread
    from google.oauth2.service_account import Credentials
    from gspread.exceptions import WorksheetNotFound, APIError
except Exception:
    gspread = None
    Credentials = None
    WorksheetNotFound = APIError = Exception  # fallbacks

# ===================== Money/Date Utilities =====================
_MONEY_TOKEN   = re.compile(r'^[\$\d,\.]+$')
_MONEY_STRICT  = re.compile(r'^\$?\d{1,3}(?:,\d{3})*(?:\.\d{2})?$|^\$?\d+(?:\.\d{2})?$')
_DATE_LIKE     = re.compile(r'^(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})$')

def _normalize_amount(txt: str) -> str:
    s = txt.strip().replace(',', '').replace(' ', '').lstrip('$')
    m = re.search(r'\d+(?:\.\d{1,2})?', s)
    return m.group(0) if m else ''

def _looks_like_amount(txt: str) -> bool:
    return bool(_MONEY_STRICT.match(txt.strip().replace(' ', '')))

def _normalize_date_guess(s: str) -> str:
    s = s.strip()
    # Fix jammed mmdd/year (e.g., "1102/2019" -> "11/02/2019")
    m = re.match(r'^(\d{4})/(\d{2,4})$', s)
    if m:
        mmdd, yy = m.group(1), m.group(2)
        if len(mmdd) == 4:
            mm, dd = mmdd[:2], mmdd[2:]
            s = f"{mm}/{dd}/{yy}"

    # try common formats
    for fmt in ('%m/%d/%Y','%d/%m/%Y','%Y-%m-%d','%Y/%m/%d','%m-%d-%Y','%d-%m-%Y','%m/%d/%y','%d/%m/%y'):
        try:
            dt = datetime.strptime(s, fmt)
            # clamp year to something plausible
            now_y = datetime.now().year
            if 1990 <= dt.year <= now_y + 1:
                return dt.strftime('%Y-%m-%d')
        except Exception:
            continue
    return s  # fallback as-is

def _collect_contiguous_amount(words, j):
    tokens, k = [], j
    while k < len(words) and _MONEY_TOKEN.match(words[k]):
        tokens.append(words[k]); k += 1
    return _normalize_amount(''.join(tokens)) if tokens else ''

def _pick_right_of_keyword(words, boxes, keyset, accept_fn, max_lookahead=30, same_line_tol=25):
    """Find acceptable value to the right of a keyword on the same row."""
    yc = [ (b[1]+b[3])//2 for b in boxes ]
    for i, w in enumerate(words):
        wl = w.lower().rstrip(':')
        if wl in keyset:
            y = yc[i]; best, best_dx = None, 10**9
            for j in range(i+1, min(i+1+max_lookahead, len(words))):
                if abs(yc[j]-y) <= same_line_tol:
                    cand = words[j]
                    if accept_fn(cand):
                        dx = boxes[j][0] - boxes[i][2]
                        if 0 <= dx < best_dx:
                            best, best_dx = cand, dx
                    if accept_fn == _looks_like_amount:
                        amt = _collect_contiguous_amount(words, j)
                        if amt:
                            dx = boxes[j][0] - boxes[i][2]
                            if 0 <= dx < best_dx:
                                best, best_dx = amt, dx
            if best:
                return best
    return None

def _largest_amount(words):
    vals = []
    for w in words:
        if _looks_like_amount(w):
            try:
                v = float(_normalize_amount(w))
                # Heuristic: ignore lone small integers like "1" that often come from QTY/indices
                if ('$' in w) or ('.' in w) or v >= 10.0:
                    vals.append(v)
            except Exception:
                pass
    return f"{max(vals):.2f}" if vals else None

# ===================== Heuristic Post-processing =====================
def postprocess_fields(words, boxes, fields):
    if 'TOTAL' in fields:
        fields['TOTAL'] = _normalize_amount(fields['TOTAL'])
    if not fields.get('TOTAL') or not _looks_like_amount(fields['TOTAL']):
        by_kw = _pick_right_of_keyword(
            words, boxes,
            keyset={'total','amount due','grand total','balance due','invoice total'},
            accept_fn=_looks_like_amount
        )
        if by_kw:
            fields['TOTAL'] = _normalize_amount(by_kw)
        elif not fields.get('TOTAL'):
            mx = _largest_amount(words)
            if mx: fields['TOTAL'] = mx

    date_kw = _pick_right_of_keyword(
        words, boxes,
        keyset={'invoice date','date','issue date'},
        accept_fn=lambda s: bool(_DATE_LIKE.match(s))
    )
    if date_kw:
        fields['DATE'] = _normalize_date_guess(date_kw)
    elif 'DATE' in fields:
        fields['DATE'] = _normalize_date_guess(fields['DATE'])
    return fields

# ===================== Metadata & Items Extractors =====================
_ALNUM_ID = re.compile(r'^[A-Za-z0-9\-_/]+$')
_INVNUM   = re.compile(r'^(?=.*\d)[A-Za-z0-9\-_\/]+$')  # must contain a digit

def extract_company(words, boxes):
    top = [(i, (b[0]+b[2])//2, (b[1]+b[3])//2) for i,b in enumerate(boxes) if b[1] < 250]
    if not top: return ''
    top.sort(key=lambda t: (t[2], t[1]))
    lines, curr_y, buf = [], None, []
    for i, xc, yc in top:
        if curr_y is None or abs(yc - curr_y) <= 20:
            buf.append(words[i]); curr_y = yc if curr_y is None else curr_y
        else:
            lines.append(' '.join(buf)); buf = [words[i]]; curr_y = yc
    if buf: lines.append(' '.join(buf))
    for ln in lines:
        if re.search(r'\b(inc|llc|ltd|ltda|pty|pvt|co\.?)\b', ln, re.I):
            return ln.strip()
    for ln in lines:
        if 'invoice' not in ln.lower() and any(c.isalpha() for c in ln):
            return ln.strip()
    return ''

_INVNUM = re.compile(r'^(?=.*\d)[A-Za-z0-9][A-Za-z0-9\-/_.]{2,}$')  # must contain a digit, min len 3

def extract_invoice_number(words, boxes):
    keys = {'invoice#','invoice #','invoice no','invoice no.','invoice number','invoice','inv','inv#','inv no','inv no.'}
    yc = [ (b[1]+b[3])//2 for b in boxes ]
    stop = {'date','due','po','#','no','number','p.o.#','p.o.','gst','tax'}

    # 1) Keyword to the right on same line
    for i, w in enumerate(words):
        wl = w.lower().rstrip(':')
        if wl in keys:
            y = yc[i]
            j_start = i + 1
            # skip separators
            while j_start < len(words) and words[j_start].lower() in {'#','no','no.',':'}:
                j_start += 1
            best, best_dx = None, 10**9
            for j in range(j_start, min(i+50, len(words))):
                if abs(yc[j]-y) > 28:  # same row
                    continue
                cand = words[j].strip()
                if cand.lower() in stop:
                    continue
                if _INVNUM.match(cand):
                    dx = boxes[j][0] - boxes[i][2]
                    if 0 <= dx < best_dx:
                        best, best_dx = cand, dx
            if best:
                return best

    # 2) Top-right fallback (common placement): pick the most “ID-like” token there
    # consider tokens in upper 35% height and right 45% width
    ids = []
    for i,(w,b) in enumerate(zip(words, boxes)):
        x0,y0,x1,y1 = b
        if y0 < 350 and x0 > 550 and _INVNUM.match(w):
            ids.append((i,w))
    # prefer tokens near the word "invoice" in Y
    if ids:
        inv_y = None
        for i,w in enumerate(words):
            if w.lower().startswith('invoice'):
                inv_y = yc[i]; break
        if inv_y is not None:
            ids.sort(key=lambda t: abs(yc[t[0]] - inv_y))
        else:
            ids.sort(key=lambda t: (-len(t[1]), t[0]))
        return ids[0][1]

    return ""


def extract_due_date(words, boxes):
    v = _pick_right_of_keyword(
        words, boxes,
        keyset={'due date','payment due'},
        accept_fn=lambda s: bool(_DATE_LIKE.match(s))
    )
    return _normalize_date_guess(v) if v else ''

def extract_items(words, boxes):
    # normalize tokens for matching (strip punctuation/colon)
    def norm(w): return re.sub(r'[^a-z0-9]+', '', w.lower())

    SYN = {
        'QTY': {'qty','quantity','q'},
        'DESCRIPTION': {'description','desc','item','details','particulars'},
        'UNIT PRICE': {'unitprice','unit','price','rate','unitrate','unit_cost','unitcost'},
        'AMOUNT': {'amount','linetotal','total','ext','extamount'}
    }

    # find header columns by first match
    header_pos = {}
    for i, w in enumerate(words):
        nw = norm(w)
        for col, keys in SYN.items():
            if nw in keys:
                header_pos.setdefault(col, i)

    # handle "UNIT PRICE" split headers (UNIT + PRICE)
    if 'UNIT PRICE' not in header_pos:
        u = p = None
        for i, w in enumerate(words):
            nw = norm(w)
            if nw == 'unit' and u is None: u = i
            if nw == 'price' and p is None: p = i
        if u is not None and p is not None:
            header_pos['UNIT PRICE'] = ('COMBO', u, p)

    if not {'DESCRIPTION','AMOUNT'} & set(header_pos.keys()):
        return []

    def xcenter(i): b = boxes[i]; return (b[0]+b[2])//2
    cols = {}
    if 'QTY' in header_pos: cols['QTY'] = xcenter(header_pos['QTY'])
    if 'DESCRIPTION' in header_pos: cols['DESCRIPTION'] = xcenter(header_pos['DESCRIPTION'])
    if 'UNIT PRICE' in header_pos:
        v = header_pos['UNIT PRICE']
        if isinstance(v, tuple): _, i1, i2 = v; cols['UNIT PRICE'] = (xcenter(i1)+xcenter(i2))//2
        else: cols['UNIT PRICE'] = xcenter(v)
    if 'AMOUNT' in header_pos: cols['AMOUNT'] = xcenter(header_pos['AMOUNT'])

    # y threshold: just below the top-most header token
    header_y = min(b for i,b in enumerate([boxes[idx] for idx in [v for v in header_pos.values() if isinstance(v,int)] ]) or [[0,0,0,0]])
    header_y = header_y[1] if isinstance(header_y, list) else 0

    # collect tokens below header; sort by (y,x)
    toks=[]
    for (w,b) in zip(words, boxes):
        yc = (b[1]+b[3])//2
        xc = (b[0]+b[2])//2
        if yc <= header_y:  # only rows below header
            continue
        toks.append((w,xc,yc,b))
    toks.sort(key=lambda t:(t[2],t[1]))

    # dynamic y-merge tolerance based on median token height
    heights = [(b[3]-b[1]) for (_,_,_,b) in toks] or [22]
    tol = max(22, int(sorted(heights)[len(heights)//2] * 1.4))

    # group into lines
    lines=[]; line=[]; last_y=None
    for t in toks:
        yc=t[2]
        if last_y is None or abs(yc-last_y) <= tol:
            line.append(t); last_y=yc if last_y is None else last_y
        else:
            lines.append(line); line=[t]; last_y=yc
    if line: lines.append(line)

    items=[]
    for ln in lines:
        if not ln: continue
        joined = ' '.join(w.lower() for (w,_,_,_) in ln)
        if any(k in joined for k in ['subtotal','tax','total','balance']):
            continue
        buckets={'QTY':[],'DESCRIPTION':[],'UNIT PRICE':[],'AMOUNT':[]}
        for (w,xc,_,_) in ln:
            nearest = min(cols.keys(), key=lambda c: abs(xc - cols[c]))
            buckets[nearest].append(w)

        qty_txt = ' '.join(buckets['QTY']).strip()
        desc    = ' '.join(buckets['DESCRIPTION']).strip()
        unit    = _normalize_amount(' '.join(buckets['UNIT PRICE']))
        amount  = _normalize_amount(' '.join(buckets['AMOUNT']))
        if not amount:
            m = re.search(r'\$?\d{1,3}(?:,\d{3})*(?:\.\d{2})?$', desc)
            if m: amount = _normalize_amount(m.group(0)); desc = desc[:m.start()].strip()

        if desc and (amount or unit):
            try: qty = int(re.sub(r'\D+', '', qty_txt)) if qty_txt else None
            except: qty = None
            items.append({"DESCRIPTION": desc, "QTY": qty, "UNIT_PRICE": unit or '', "AMOUNT": amount or ''})
    return items


# ===================== Model Init =====================
device = "cuda" if torch.cuda.is_available() else "cpu"
try:
    processor = AutoProcessor.from_pretrained(MODEL_DIR)
    model = LayoutLMv3ForTokenClassification.from_pretrained(MODEL_DIR).to(device)
    id2label = {int(k) if isinstance(k, str) else k: v for k, v in model.config.id2label.items()}
    label_set = sorted(set(id2label.values()))
except Exception as e:
    raise RuntimeError(f"Failed to load model or processor from {MODEL_DIR}: {e}")

excel_lock = threading.Lock()

# ===================== OCR =====================
def normalize_box(left, top, width, height, img_w, img_h):
    x0 = int(1000 * left / img_w); y0 = int(1000 * top / img_h)
    x1 = int(1000 * (left + width) / img_w); y1 = int(1000 * (top + height) / img_h)
    return [max(0,min(1000,x0)), max(0,min(1000,y0)), max(0,min(1000,x1)), max(0,min(1000,y1))]

def ocr_words_and_boxes(image: Image.Image):
    image = ImageOps.exif_transpose(image)
    W, H = image.size
    data = pytesseract.image_to_data(
        image,
        output_type=pytesseract.Output.DICT,
        config="--oem 1 --psm 6"
    )
    words, boxes = [], []
    n = len(data.get("text", []))
    for i in range(n):
        txt = (data["text"][i] or "").strip()
        if not txt:
            continue
        l, t = int(data["left"][i]), int(data["top"][i])
        w, h = int(data["width"][i]), int(data["height"][i])
        words.append(txt)
        boxes.append(normalize_box(l, t, w, h, W, H))
    if not words:
        raise ValueError("OCR produced no words—check Tesseract install/image quality.")
    return image, words, boxes

# ===================== BIO Aggregation =====================
def aggregate_entities(
    words: List[str],
    boxes: List[List[int]],
    token_preds: List[int],
    word_ids: List[int],
) -> Dict[str, str]:
    per_word_labels = defaultdict(list)
    for i, wid in enumerate(word_ids):
        if wid is not None:
            per_word_labels[wid].append(id2label[token_preds[i]])

    max_wid = max([w for w in word_ids if w is not None]) if any(w is not None for w in word_ids) else -1
    word_labels = [Counter(per_word_labels.get(wid, ["O"])).most_common(1)[0][0] for wid in range(max_wid + 1)]

    fields, curr_type, buffer = {}, None, []
    for w, tag in zip(words[:len(word_labels)], word_labels):
        if tag == "O":
            if curr_type:
                fields[curr_type] = (fields.get(curr_type, "") + " " + " ".join(buffer)).strip()
                curr_type, buffer = None, []
            continue
        prefix, _, ent = tag.partition("-")
        if ent == "DATE" and not any(ch.isdigit() for ch in w):
            continue
        if prefix == "B":
            if curr_type:
                fields[curr_type] = (fields.get(curr_type, "") + " " + " ".join(buffer)).strip()
            curr_type, buffer = ent, [w]
        elif prefix == "I":
            if curr_type == ent:
                buffer.append(w)
            else:
                if curr_type:
                    fields[curr_type] = (fields.get(curr_type, "") + " " + " ".join(buffer)).strip()
                curr_type, buffer = ent, [w]
    if curr_type:
        fields[curr_type] = (fields.get(curr_type, "") + " " + " ".join(buffer)).strip()

    if "TOTAL" in fields:
        tot = fields["TOTAL"].replace(",", "")
        m = re.search(r"(\d+(?:\.\d{1,2})?)", tot)
        if m:
            fields["TOTAL"] = m.group(1)
    return fields

# ===================== Excel Writer (Flat Rows) =====================
def append_items_flat_excel(meta: Dict[str, Any], items: list, excel_path: str = EXCEL_PATH):
    cols = ["COMPANY","INVOICE_NO","DATE","DUE_DATE","ITEM_DESCRIPTION","QTY","UNIT_PRICE","AMOUNT","TOTAL"]
    Path(os.path.dirname(excel_path)).mkdir(parents=True, exist_ok=True)
    with excel_lock:
        if not os.path.exists(excel_path):
            wb = Workbook(); ws = wb.active; ws.title = "rows"; ws.append(cols); wb.save(excel_path)
        wb = load_workbook(excel_path)
        ws = wb["rows"] if "rows" in wb.sheetnames else wb.create_sheet("rows")
        if ws.max_row == 1: ws.append(cols)
        if not items:
            ws.append([meta.get("COMPANY",""), meta.get("INVOICE_NO",""), meta.get("DATE",""), meta.get("DUE_DATE",""),
                       "", "", "", "", meta.get("TOTAL","")])
        else:
            for it in items:
                ws.append([
                    meta.get("COMPANY",""),
                    meta.get("INVOICE_NO",""),
                    meta.get("DATE",""),
                    meta.get("DUE_DATE",""),
                    it.get("DESCRIPTION",""),
                    it.get("QTY",""),
                    it.get("UNIT_PRICE",""),
                    it.get("AMOUNT",""),
                    meta.get("TOTAL",""),
                ])
        wb.save(excel_path)
        logger.info(f"✔ wrote {(len(items) or 1)} row(s) to Excel {excel_path} [rows]")

# ===================== Google Sheets Writer (Flat Rows) =====================
def _validate_sa_json(data: dict) -> str:
    t = data.get("type")
    if t != "service_account":
        raise RuntimeError(f"Provided JSON is not a service account key (type='{t}'). Expected type='service_account'.")
    email = data.get("client_email")
    if not email:
        raise RuntimeError("Service account JSON missing 'client_email'.")
    return email

def _load_sa_json() -> tuple[dict, str]:
    """
    Returns (json_dict, source), where source is 'inline_json' or the file path.
    Prefers GOOGLE_CRED_JSON if both are set.
    """
    both_set = bool(GOOGLE_CRED_JSON) and bool(GOOGLE_APPLICATION_CREDENTIALS)
    if GOOGLE_CRED_JSON:
        try:
            data = json.loads(GOOGLE_CRED_JSON)
        except Exception as e:
            raise RuntimeError(f"GOOGLE_CRED_JSON not valid JSON: {e}")
        # still fine if both_set; we intentionally prefer inline JSON
        return data, "inline_json"
    if GOOGLE_APPLICATION_CREDENTIALS:
        if not os.path.exists(GOOGLE_APPLICATION_CREDENTIALS):
            raise RuntimeError(f"GOOGLE_APPLICATION_CREDENTIALS not found: {GOOGLE_APPLICATION_CREDENTIALS}")
        with open(GOOGLE_APPLICATION_CREDENTIALS, "r") as f:
            data = json.load(f)
        return data, GOOGLE_APPLICATION_CREDENTIALS
    raise RuntimeError("Provide GOOGLE_APPLICATION_CREDENTIALS or GOOGLE_CRED_JSON for Google Sheets.")

def _resolve_gsheets_credentials():
    """
    Validates SA JSON and returns (cred_path, client_email, source).
    """
    data, source = _load_sa_json()
    email = _validate_sa_json(data)
    if source == "inline_json":
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
        tmp.write(json.dumps(data).encode()); tmp.flush()
        return tmp.name, email, source
    return source, email, source  # source is a file path here

def append_items_gsheets(meta: Dict[str, Any], items: list):
    if gspread is None or Credentials is None:
        raise RuntimeError("gspread/google-auth not installed")
    if not GSHEETS_ID:
        raise RuntimeError("GSHEETS_ID not set")
    cred_path, _, _ = _resolve_gsheets_credentials()
    creds = Credentials.from_service_account_file(cred_path, scopes=GSHEETS_SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(GSHEETS_ID)
    try:
        ws = sh.worksheet("rows")
    except WorksheetNotFound:
        ws = sh.add_worksheet(title="rows", rows=1, cols=9)
        ws.append_row(["COMPANY","INVOICE_NO","DATE","DUE_DATE","ITEM_DESCRIPTION","QTY","UNIT_PRICE","AMOUNT","TOTAL"])

    rows = []
    if not items:
        rows.append([meta.get("COMPANY",""), meta.get("INVOICE_NO",""), meta.get("DATE",""), meta.get("DUE_DATE",""),
                     "", "", "", "", meta.get("TOTAL","")])
    else:
        for it in items:
            rows.append([
                meta.get("COMPANY",""),
                meta.get("INVOICE_NO",""),
                meta.get("DATE",""),
                meta.get("DUE_DATE",""),
                it.get("DESCRIPTION",""),
                it.get("QTY",""),
                it.get("UNIT_PRICE",""),
                it.get("AMOUNT",""),
                meta.get("TOTAL",""),
            ])
    try:
        ws.append_rows(rows, value_input_option="USER_ENTERED")
    except APIError as e:
        # Bubble a friendlier hint for permissions
        raise RuntimeError(f"Google Sheets API error (likely share/permission issue): {e}")
    logger.info(f"✔ wrote {len(rows)} row(s) to Google Sheets {GSHEETS_ID}/rows")

def append_items(meta: Dict[str, Any], items: list):
    wrote, errors = [], []
    if WRITE_SINK in ("excel", "both"):
        try:
            append_items_flat_excel(meta, items, EXCEL_PATH)
            wrote.append(f"excel:{EXCEL_PATH}")
        except Exception as e:
            errors.append(f"excel:{e}")
    if WRITE_SINK in ("gsheets", "both"):
        try:
            append_items_gsheets(meta, items)
            wrote.append("gsheets")
        except Exception as e:
            logger.exception("Google Sheets write failed")
            errors.append(f"gsheets:{e}")
    return wrote, errors

# ===================== Routes =====================
@app.get("/health")
def health():
    return {"status": "ok", "device": device, "labels": label_set, "sink": WRITE_SINK}

@app.get("/where")
def where():
    diag = {
        "write_sink": WRITE_SINK,
        "excel_path": EXCEL_PATH,
        "gsheets_id": GSHEETS_ID or "",
        "gspread_installed": bool(gspread),
        "both_env_set": bool(GOOGLE_CRED_JSON) and bool(GOOGLE_APPLICATION_CREDENTIALS),
        "creds_source": "",
        "creds_type": "",
        "creds_path_exists": False,
        "service_account_email": "",
        "share_hint": "",
    }
    try:
        data, source = _load_sa_json()
        diag["creds_source"] = source
        diag["creds_type"] = data.get("type", "")
        if source != "inline_json":
            diag["creds_path_exists"] = os.path.exists(source)
        email = _validate_sa_json(data)
        diag["service_account_email"] = email
        diag["share_hint"] = f"Share the Sheet with {email} as Editor."
    except Exception as e:
        diag["creds_error"] = str(e)
    return diag

@app.get("/gsheets/probe")
def gsheets_probe():
    try:
        cred_path, email, source = _resolve_gsheets_credentials()
        creds = Credentials.from_service_account_file(cred_path, scopes=GSHEETS_SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(GSHEETS_ID)
        tabs = [ws.title for ws in sh.worksheets()]
        return {"ok": True, "service_account_email": email, "creds_source": source, "tabs": tabs}
    except Exception as e:
        return {"ok": False, "error": str(e), "gsheets_id": GSHEETS_ID}

@app.get("/download")
def download_excel():
    if not os.path.exists(EXCEL_PATH):
        return PlainTextResponse("No Excel file yet.", status_code=404)
    return FileResponse(
        EXCEL_PATH,
        filename=os.path.basename(EXCEL_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/predict")
async def predict(file: UploadFile = File(...)):
    # 1) Read image
    try:
        content = await file.read()
        image = Image.open(io.BytesIO(content)).convert("RGB")
    except Exception:
        raise HTTPException(status_code=400, detail="Unsupported image or corrupted file.")

    # 2) OCR
    try:
        image, words, boxes = ocr_words_and_boxes(image)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"OCR failed: {e}")

    # 3) Encode → model
    try:
        encoding = processor(
            images=image,
            text=words,
            boxes=boxes,
            return_tensors="pt",
            truncation=True,
            padding="max_length",
            max_length=MAX_LENGTH,
        )
        word_ids = encoding.word_ids(batch_index=0)
        inputs = {k: (v.to(device) if hasattr(v, "to") else v) for k, v in encoding.items()}
        with torch.no_grad():
            outputs = model(**inputs)
        preds = outputs.logits.argmax(-1)[0].tolist()
        fields = aggregate_entities(words, boxes, preds, word_ids)
        fields = postprocess_fields(words, boxes, fields)
    except Exception as e:
        logger.exception("Inference failed")
        raise HTTPException(status_code=500, detail=f"Inference failed: {e}")

    # 4) Metadata + items
    try:
        meta = {
            "COMPANY": extract_company(words, boxes) or fields.get("COMPANY",""),
            "INVOICE_NO": extract_invoice_number(words, boxes) or "",
            "DATE": fields.get("DATE",""),
            "DUE_DATE": extract_due_date(words, boxes) or "",
            "TOTAL": fields.get("TOTAL",""),
        }
        items = extract_items(words, boxes)

        # --- TOTAL fix-up: prefer item-derived totals if model/heuristics picked a tiny number ---
        def _sum_items(itms):
            vals=[]
            for it in itms:
                try:
                    v = float(it.get("AMOUNT") or it.get("UNIT_PRICE") or "0")
                    if v>0: vals.append(v)
                except: pass
            return sum(vals) if vals else 0.0

        if items:
            items_total = _sum_items(items)
            try:
                t = float(meta.get("TOTAL") or 0)
            except:
                t = 0.0
            # if current total is missing, < 10, or far below item sum, trust items
            if (t == 0.0) or (t < 10 and items_total > t) or (items_total > 0 and items_total > t*0.7):
                meta["TOTAL"] = f"{items_total:.2f}"
    except Exception as e:
        logger.exception("Postprocess failed")
        raise HTTPException(status_code=500, detail=f"Postprocess failed: {e}")


    # 5) Persist
    wrote, errors = append_items(meta, items)
    return JSONResponse({"fields": meta, "items": items, "wrote": wrote, "errors": errors})
